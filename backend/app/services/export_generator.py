"""
Export generator service for creating Word and Excel reports.
Supports multiple export formats and customization options for accreditation purposes.
"""

import pandas as pd
import numpy as np
from datetime import datetime
from typing import List, Dict, Optional, Tuple
from io import BytesIO
import os
from pathlib import Path
import json

from app.models.export import ExportFormat, ExportType, ExportJob
from app.models.publication import Publication, PublicationType
from app.models.faculty import Faculty
from app.models.plagiarism import PlagiarismScan


class ExportGenerator:
    """Service for generating Word and Excel export reports."""

    def __init__(self):
        self.citation_styles = {
            'APA': self._format_apa_citation,
            'MLA': self._format_mla_citation,
            'IEEE': self._format_ieee_citation,
            'Chicago': self._format_chicago_citation
        }

    async def generate_export(
        self,
        export_job: ExportJob,
        publications: List[Publication],
        faculty: Optional[List[Faculty]] = None,
        plagiarism_scans: Optional[List[PlagiarismScan]] = None
    ) -> bytes:
        """
        Generate export file based on job configuration.

        Args:
            export_job: Export job configuration
            publications: List of publications to include
            faculty: List of faculty (optional)
            plagiarism_scans: List of plagiarism scans (optional)

        Returns:
            File content as bytes
        """
        try:
            if export_job.export_format == ExportFormat.EXCEL:
                return await self._generate_excel_export(export_job, publications, faculty, plagiarism_scans)
            elif export_job.export_format == ExportFormat.WORD:
                return await self._generate_word_export(export_job, publications, faculty, plagiarism_scans)
            elif export_job.export_format == ExportFormat.CSV:
                return await self._generate_csv_export(export_job, publications)
            elif export_job.export_format == ExportFormat.JSON:
                return await self._generate_json_export(export_job, publications, faculty, plagiarism_scans)
            else:
                raise ValueError(f"Unsupported export format: {export_job.export_format}")

        except Exception as e:
            raise Exception(f"Export generation failed: {str(e)}")

    async def _generate_excel_export(
        self,
        export_job: ExportJob,
        publications: List[Publication],
        faculty: Optional[List[Faculty]],
        plagiarism_scans: Optional[List[PlagiarismScan]]
    ) -> bytes:
        """Generate Excel export file."""
        # Import openpyxl here to avoid import issues if not available
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            raise Exception("openpyxl library is required for Excel export")

        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Filter publications based on export criteria
        filtered_publications = await self._filter_publications(publications, export_job)

        # Create different sheets based on export type
        if export_job.export_type == ExportType.YEAR_WISE:
            await self._create_year_wise_sheets(wb, filtered_publications, export_job)
        elif export_job.export_type == ExportType.TYPE_WISE:
            await self._create_type_wise_sheets(wb, filtered_publications, export_job)
        elif export_job.export_type == ExportType.FACULTY_WISE:
            await self._create_faculty_wise_sheets(wb, filtered_publications, faculty, export_job)
        elif export_job.export_type == ExportType.DEPARTMENT_WISE:
            await self._create_department_wise_sheets(wb, filtered_publications, faculty, export_job)
        elif export_job.export_type == ExportType.PLAGIARISM_REPORT:
            await self._create_plagiarism_report_sheets(wb, filtered_publications, plagiarism_scans, export_job)
        else:
            # Default: Create summary sheet
            await self._create_summary_sheet(wb, filtered_publications, faculty, plagiarism_scans, export_job)

        # Save to bytes
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        return excel_buffer.getvalue()

    async def _generate_word_export(
        self,
        export_job: ExportJob,
        publications: List[Publication],
        faculty: Optional[List[Faculty]],
        plagiarism_scans: Optional[List[PlagiarismScan]]
    ) -> bytes:
        """Generate Word export file."""
        # For now, create a simple text-based export since python-docx might not be available
        # In a real implementation, you would use python-docx library

        filtered_publications = await self._filter_publications(publications, export_job)

        # Create content
        content = []
        content.append(export_job.title)
        content.append("=" * len(export_job.title))
        content.append(f"Export Date: {datetime.now().strftime('%B %d, %Y')}")
        content.append(f"Total Publications: {len(filtered_publications)}")
        content.append("")

        # Group by type
        type_groups = {}
        for pub in filtered_publications:
            pub_type = pub.publication_type.value.replace('_', ' ').title()
            if pub_type not in type_groups:
                type_groups[pub_type] = []
            type_groups[pub_type].append(pub)

        for pub_type, pubs in sorted(type_groups.items()):
            content.append(f"\n{pub_type.upper()}")
            content.append("-" * len(pub_type))

            for pub in pubs:
                content.append(f"\n• {pub.title}")
                if pub.authors:
                    content.append(f"  Authors: {', '.join(pub.authors)}")
                content.append(f"  Year: {pub.publication_year}")
                if pub.venue_name:
                    content.append(f"  Venue: {pub.venue_name}")
                if pub.doi:
                    content.append(f"  DOI: {pub.doi}")

        # Convert to bytes
        text_content = "\n".join(content)
        return text_content.encode('utf-8')

    async def _generate_csv_export(
        self,
        export_job: ExportJob,
        publications: List[Publication]
    ) -> bytes:
        """Generate CSV export file."""
        filtered_publications = await self._filter_publications(publications, export_job)

        # Create DataFrame
        data = []
        for pub in filtered_publications:
            row = {
                'Title': pub.title,
                'Authors': '; '.join(pub.authors) if pub.authors else '',
                'Year': pub.publication_year,
                'Type': pub.publication_type.value,
                'Venue': pub.venue_name or '',
                'DOI': pub.doi or '',
                'Citations': pub.citation_count or 0,
                'Impact Factor': pub.impact_factor or 0
            }
            data.append(row)

        df = pd.DataFrame(data)

        # Save to CSV
        csv_buffer = BytesIO()
        df.to_csv(csv_buffer, index=False, encoding='utf-8-sig')
        csv_buffer.seek(0)

        return csv_buffer.getvalue()

    async def _generate_json_export(
        self,
        export_job: ExportJob,
        publications: List[Publication],
        faculty: Optional[List[Faculty]],
        plagiarism_scans: Optional[List[PlagiarismScan]]
    ) -> bytes:
        """Generate JSON export file."""
        filtered_publications = await self._filter_publications(publications, export_job)

        # Create export data structure
        export_data = {
            'metadata': {
                'export_date': datetime.now().isoformat(),
                'export_type': export_job.export_type.value,
                'title': export_job.title,
                'total_publications': len(filtered_publications),
                'date_range': {
                    'from': export_job.year_from,
                    'to': export_job.year_to
                }
            },
            'publications': []
        }

        for pub in filtered_publications:
            pub_data = {
                'id': pub.id,
                'title': pub.title,
                'abstract': pub.abstract,
                'authors': pub.authors,
                'author_count': pub.author_count,
                'publication_year': pub.publication_year,
                'publication_type': pub.publication_type.value,
                'venue_name': pub.venue_name,
                'publisher': pub.publisher,
                'doi': pub.doi,
                'url': pub.url,
                'citation_count': pub.citation_count,
                'impact_factor': pub.impact_factor,
                'relevance_score': pub.relevance_score,
                'confidence_score': pub.confidence_score,
                'source_database': pub.source_database
            }
            export_data['publications'].append(pub_data)

        # Add faculty information if available
        if faculty:
            export_data['faculty'] = [
                {
                    'id': f.id,
                    'name': f.name,
                    'department': f.department,
                    'email': f.email,
                    'title': f.title
                }
                for f in faculty
            ]

        # Add plagiarism scan results if requested
        if export_job.include_plagiarism_data and plagiarism_scans:
            export_data['plagiarism_scans'] = []
            for scan in plagiarism_scans:
                scan_data = {
                    'publication_id': scan.publication_id,
                    'overall_similarity_score': scan.overall_similarity_score,
                    'severity_level': scan.severity_level,
                    'scan_date': scan.scan_date.isoformat(),
                    'status': scan.status
                }
                export_data['plagiarism_scans'].append(scan_data)

        # Serialize to JSON
        json_str = json.dumps(export_data, indent=2, ensure_ascii=False, default=str)
        return json_str.encode('utf-8')

    async def _filter_publications(
        self,
        publications: List[Publication],
        export_job: ExportJob
    ) -> List[Publication]:
        """Filter publications based on export job criteria."""
        filtered = []

        for pub in publications:
            # Year range filter
            if export_job.year_from and pub.publication_year < export_job.year_from:
                continue
            if export_job.year_to and pub.publication_year > export_job.year_to:
                continue

            # Publication type filter
            if export_job.publication_types:
                if pub.publication_type.value not in export_job.publication_types:
                    continue

            # Faculty filter
            if export_job.faculty_ids and pub.faculty_id not in export_job.faculty_ids:
                continue

            filtered.append(pub)

        return filtered

    async def _create_year_wise_sheets(
        self,
        wb,
        publications: List[Publication],
        export_job: ExportJob
    ):
        """Create separate sheets for each year."""
        # Group publications by year
        years = {}
        for pub in publications:
            year = pub.publication_year
            if year not in years:
                years[year] = []
            years[year].append(pub)

        # Create sheet for each year
        for year, year_pubs in sorted(years.items()):
            ws = wb.create_sheet(title=str(year))
            await self._populate_publication_sheet(ws, year_pubs, export_job)

    async def _create_type_wise_sheets(
        self,
        wb,
        publications: List[Publication],
        export_job: ExportJob
    ):
        """Create separate sheets for each publication type."""
        # Group publications by type
        types = {}
        for pub in publications:
            pub_type = pub.publication_type.value.replace('_', ' ').title()
            if pub_type not in types:
                types[pub_type] = []
            types[pub_type].append(pub)

        # Create sheet for each type
        for pub_type, type_pubs in sorted(types.items()):
            ws = wb.create_sheet(title=pub_type)
            await self._populate_publication_sheet(ws, type_pubs, export_job)

    async def _create_faculty_wise_sheets(
        self,
        wb,
        publications: List[Publication],
        faculty: Optional[List[Faculty]],
        export_job: ExportJob
    ):
        """Create separate sheets for each faculty member."""
        if not faculty:
            # Group by faculty_id if no faculty list provided
            faculty_groups = {}
            for pub in publications:
                faculty_id = pub.faculty_id
                if faculty_id not in faculty_groups:
                    faculty_groups[faculty_id] = []
                faculty_groups[faculty_id].append(pub)

            for faculty_id, faculty_pubs in faculty_groups.items():
                ws = wb.create_sheet(title=f"Faculty_{faculty_id}")
                await self._populate_publication_sheet(ws, faculty_pubs, export_job)
        else:
            # Create faculty mapping
            faculty_map = {f.id: f for f in faculty}

            # Group publications by faculty
            faculty_groups = {}
            for pub in publications:
                if pub.faculty_id in faculty_map:
                    faculty_name = faculty_map[pub.faculty_id].name
                    if faculty_name not in faculty_groups:
                        faculty_groups[faculty_name] = []
                    faculty_groups[faculty_name].append(pub)

            # Create sheet for each faculty member
            for faculty_name, faculty_pubs in sorted(faculty_groups.items()):
                # Clean sheet name (Excel has limitations)
                sheet_name = faculty_name[:30]  # Excel sheet name limit
                sheet_name = ''.join(c for c in sheet_name if c.isalnum() or c in (' ', '-', '_'))

                ws = wb.create_sheet(title=sheet_name)
                await self._populate_publication_sheet(ws, faculty_pubs, export_job)

    async def _create_department_wise_sheets(
        self,
        wb,
        publications: List[Publication],
        faculty: Optional[List[Faculty]],
        export_job: ExportJob
    ):
        """Create separate sheets for each department."""
        if not faculty:
            return

        # Create faculty to department mapping
        faculty_depts = {f.id: f.department for f in faculty}

        # Group publications by department
        dept_groups = {}
        for pub in publications:
            if pub.faculty_id in faculty_depts:
                dept = faculty_depts[pub.faculty_id]
                if dept not in dept_groups:
                    dept_groups[dept] = []
                dept_groups[dept].append(pub)

        # Create sheet for each department
        for dept, dept_pubs in sorted(dept_groups.items()):
            # Clean sheet name
            sheet_name = dept[:30]
            sheet_name = ''.join(c for c in sheet_name if c.isalnum() or c in (' ', '-', '_'))

            ws = wb.create_sheet(title=sheet_name)
            await self._populate_publication_sheet(ws, dept_pubs, export_job)

    async def _create_summary_sheet(
        self,
        wb,
        publications: List[Publication],
        faculty: Optional[List[Faculty]],
        plagiarism_scans: Optional[List[PlagiarismScan]],
        export_job: ExportJob
    ):
        """Create summary sheet with all publications."""
        ws = wb.create_sheet(title="Summary")
        await self._populate_publication_sheet(ws, publications, export_job)

        # Add statistics sheet
        stats_ws = wb.create_sheet(title="Statistics")
        await self._populate_statistics_sheet(stats_ws, publications, faculty, plagiarism_scans)

    async def _populate_publication_sheet(
        self,
        ws,
        publications: List[Publication],
        export_job: ExportJob
    ):
        """Populate a worksheet with publication data."""
        try:
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            # If openpyxl styles not available, create basic sheet
            pass

        # Define headers based on export configuration
        headers = ['Title', 'Authors', 'Year', 'Type', 'Venue']

        if export_job.include_citations:
            headers.extend(['Citations'])

        if export_job.include_impact_factors:
            headers.extend(['Impact Factor'])

        if export_job.include_plagiarism_data:
            headers.extend(['Plagiarism Score', 'Plagiarism Status'])

        headers.extend(['DOI', 'URL'])

        # Add headers
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        # Add data rows
        for row_num, pub in enumerate(publications, 2):
            col_num = 1

            # Title
            ws.cell(row=row_num, column=col_num, value=pub.title)
            col_num += 1

            # Authors
            authors = '; '.join(pub.authors) if pub.authors else ''
            ws.cell(row=row_num, column=col_num, value=authors)
            col_num += 1

            # Year
            ws.cell(row=row_num, column=col_num, value=pub.publication_year)
            col_num += 1

            # Type
            ws.cell(row=row_num, column=col_num, value=pub.publication_type.value.replace('_', ' ').title())
            col_num += 1

            # Venue
            ws.cell(row=row_num, column=col_num, value=pub.venue_name or '')
            col_num += 1

            # Citations
            if export_job.include_citations:
                ws.cell(row=row_num, column=col_num, value=pub.citation_count or 0)
                col_num += 1

            # Impact Factor
            if export_job.include_impact_factors:
                ws.cell(row=row_num, column=col_num, value=pub.impact_factor or 0)
                col_num += 1

            # Plagiarism data (placeholder - would need scan data)
            if export_job.include_plagiarism_data:
                ws.cell(row=row_num, column=col_num, value="")  # Similarity score
                col_num += 1
                ws.cell(row=row_num, column=col_num, value="")  # Status
                col_num += 1

            # DOI
            ws.cell(row=row_num, column=col_num, value=pub.doi or '')
            col_num += 1

            # URL
            ws.cell(row=row_num, column=col_num, value=pub.url or '')

    async def _populate_statistics_sheet(
        self,
        ws,
        publications: List[Publication],
        faculty: Optional[List[Faculty]],
        plagiarism_scans: Optional[List[PlagiarismScan]]
    ):
        """Populate statistics worksheet."""
        row = 1

        # General statistics
        ws.cell(row=row, column=1, value="General Statistics")
        row += 1

        ws.cell(row=row, column=1, value="Total Publications")
        ws.cell(row=row, column=2, value=len(publications))
        row += 1

        # Publication type breakdown
        type_counts = {}
        for pub in publications:
            pub_type = pub.publication_type.value.replace('_', ' ').title()
            type_counts[pub_type] = type_counts.get(pub_type, 0) + 1

        row += 1
        ws.cell(row=row, column=1, value="Publications by Type")
        row += 1

        for pub_type, count in sorted(type_counts.items()):
            ws.cell(row=row, column=1, value=pub_type)
            ws.cell(row=row, column=2, value=count)
            row += 1

        # Year breakdown
        year_counts = {}
        for pub in publications:
            year = pub.publication_year
            year_counts[year] = year_counts.get(year, 0) + 1

        row += 1
        ws.cell(row=row, column=1, value="Publications by Year")
        row += 1

        for year in sorted(year_counts.keys()):
            ws.cell(row=row, column=1, value=str(year))
            ws.cell(row=row, column=2, value=year_counts[year])
            row += 1

        # Faculty statistics
        if faculty:
            row += 1
            ws.cell(row=row, column=1, value="Faculty Statistics")
            row += 1

            ws.cell(row=row, column=1, value="Total Faculty")
            ws.cell(row=row, column=2, value=len(faculty))
            row += 1

            avg_pubs = len(publications) / len(faculty) if faculty else 0
            ws.cell(row=row, column=1, value="Average Publications per Faculty")
            ws.cell(row=row, column=2, value=f"{avg_pubs:.1f}")
            row += 1

    async def _create_plagiarism_report_sheets(
        self,
        wb,
        publications: List[Publication],
        plagiarism_scans: Optional[List[PlagiarismScan]],
        export_job: ExportJob
    ):
        """Create plagiarism report sheets."""
        # Main plagiarism summary sheet
        ws = wb.create_sheet(title="Plagiarism Summary")

        # Headers
        headers = ['Publication Title', 'Similarity Score', 'Severity', 'Scan Date', 'Status']
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        # Data
        if plagiarism_scans:
            row = 2
            for scan in plagiarism_scans:
                # Find corresponding publication
                pub = next((p for p in publications if p.id == scan.publication_id), None)
                if pub:
                    ws.cell(row=row, column=1, value=pub.title)
                    ws.cell(row=row, column=2, value=scan.overall_similarity_score)
                    ws.cell(row=row, column=3, value=scan.severity_level)
                    ws.cell(row=row, column=4, value=scan.scan_date.strftime("%Y-%m-%d"))
                    ws.cell(row=row, column=5, value=scan.status)
                    row += 1

    def _format_apa_citation(self, publication: Publication) -> str:
        """Format publication in APA citation style."""
        authors = ", ".join(publication.authors[:2]) if publication.authors else "Unknown"
        if len(publication.authors) > 2:
            authors += f", et al."

        year = publication.publication_year
        title = publication.title
        venue = publication.venue_name or ""

        citation = f"{authors} ({year}). {title}."
        if venue:
            citation += f" {venue}."

        if publication.doi:
            citation += f" https://doi.org/{publication.doi}"

        return citation

    def _format_mla_citation(self, publication: Publication) -> str:
        """Format publication in MLA citation style."""
        authors = ", ".join(publication.authors[:2]) if publication.authors else "Unknown"
        if len(publication.authors) > 2:
            authors += f", et al."

        title = publication.title
        venue = publication.venue_name or ""
        year = publication.publication_year

        citation = f'{authors}. "{title}."'
        if venue:
            citation += f" {venue},"
        citation += f" {year}."

        if publication.doi:
            citation += f" https://doi.org/{publication.doi}"

        return citation

    def _format_ieee_citation(self, publication: Publication) -> str:
        """Format publication in IEEE citation style."""
        authors = ", ".join(publication.authors[:1]) if publication.authors else "Unknown"
        if len(publication.authors) == 2:
            authors += f" and {publication.authors[1]}"
        elif len(publication.authors) > 2:
            authors += f" et al."

        title = publication.title
        venue = publication.venue_name or ""
        year = publication.publication_year

        citation = f"{authors}, \"{title},\""
        if venue:
            citation += f" {venue},"
        citation += f" {year}."

        if publication.doi:
            citation += f" https://doi.org/{publication.doi}"

        return citation

    def _format_chicago_citation(self, publication: Publication) -> str:
        """Format publication in Chicago citation style."""
        authors = ", ".join(publication.authors[:2]) if publication.authors else "Unknown"
        if len(publication.authors) > 2:
            authors += f", et al."

        year = publication.publication_year
        title = publication.title
        venue = publication.venue_name or ""

        citation = f"{authors}. {year}. \"{title}.\""
        if venue:
            citation += f" {venue}."

        if publication.doi:
            citation += f" https://doi.org/{publication.doi}"

        return citation